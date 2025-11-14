import streamlit as st
import time
import zipfile
from io import BytesIO
import pdfplumber
import pandas as pd
import re

# Afficher un loader au démarrage
with st.spinner("Chargement de l'application..."):
    time.sleep(0.5)

def extraire_info_pdf(pdf_bytes):
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            # Lire la première page
            page = pdf.pages[0]
            text = page.extract_text()
            
            # Extraire Date du paiement (après "Journée du ")
            date_paiement = ''
            match = re.search(r'Journée du\s+(\d{2}/\d{2}/\d{4})', text)
            if match:
                date_paiement = match.group(1)
            
            # Extraire Matricule (après "Matricule : ")
            matricule = ''
            match = re.search(r'Matricule\s*:\s*(\S+)', text)
            if match:
                matricule = match.group(1)
            
            # Extraire Nom du bénéficiaire (après "Bénéficiaire : ")
            nom_beneficiaire = ''
            match = re.search(r'Bénéficiaire\s*:\s*([^\n]+)', text)
            if match:
                nom_beneficiaire = match.group(1).strip()
            
            lignes = []
            
            # Pattern 1 : Lignes du tableau avec dates
            # Gère les espaces dans les montants (ex: 2 251,00)
            # Format: dd/mm/yyyy au dd/mm/yyyy Nature Quantité PrixUnitaire MontantRemboursé
            pattern1 = r'(\d{2}/\d{2}/\d{4})\s+au\s+(\d{2}/\d{2}/\d{4})\s+([A-Z\.\s]+?)\s+(\d+)\s+(-?[\d\s]+[,\.]\d{2})\s*€\s+(-?[\d\s]+[,\.]\d{2})\s*€'
            
            for match in re.finditer(pattern1, text):
                # Nettoyer les montants en enlevant les espaces
                prix_unitaire = match.group(5).replace(' ', '')
                montant_rembourse = match.group(6).replace(' ', '')
                
                lignes.append({
                    'Type': 'Montant brut',
                    'Date du paiement': date_paiement,
                    'Matricule': matricule,
                    'Nom du bénéficiaire': nom_beneficiaire,
                    'Date du': match.group(1),
                    'Date de fin': match.group(2),
                    'Nature de la prestation': match.group(3).strip(),
                    'Quantité': match.group(4),
                    'Montant brut': montant_rembourse,
                    'Montant net': 0
                })
            
            # Pattern 2 : Lignes "Total :"
            # Format: Total : 1 234,56 € ou Total : -1 234,56 €
            pattern2 = r'Total\s*:\s*(-?[\d\s]+[,\.]\d{2})\s*€'
            
            for match in re.finditer(pattern2, text):
                # Nettoyer le montant en enlevant les espaces
                montant_total = match.group(1).replace(' ', '')
                
                lignes.append({
                    'Type': 'Montant net',
                    'Date du paiement': date_paiement,
                    'Matricule': matricule,
                    'Nom du bénéficiaire': nom_beneficiaire,
                    'Date du': '',
                    'Date de fin': '',
                    'Nature de la prestation': 'Total',
                    'Quantité': 0,
                    'Montant brut': 0,
                    'Montant net': montant_total
                })
            
            return lignes if lignes else None
    except Exception as e:
        st.error(f"Erreur lors de l'extraction : {str(e)}")
        return None

# Titre
st.title("Bonjour Elise :)")

# Texte explicatif
st.write("Télécharge le dossier .zip avec l'ensemble des bordereaux d'arrêts maladies")

# Input de téléchargement
uploaded_file = st.file_uploader("Choisir un fichier", type=['zip'])

# Message de confirmation si un fichier est téléchargé
if uploaded_file is not None:
    st.success("Parfait 🥳")
    time.sleep(2)
    
    with st.spinner("Extraction en cours..."):
        # Lire le fichier zip
        with zipfile.ZipFile(BytesIO(uploaded_file.read())) as zip_file:
            # Récupérer la liste des noms de fichiers
            file_names = zip_file.namelist()
        
        time.sleep(1)  # Simulation du temps d'extraction
    
    # Afficher la liste des fichiers avec icônes
    st.write("**Fichiers trouvés dans le dossier :**")
    pdf_files = []
    for file_name in file_names:
        if not file_name.endswith('/'):  # Ignorer les dossiers
            if file_name.lower().endswith('.pdf'):
                st.write(f"{file_name} -- ✅ OK")
                pdf_files.append(file_name)
            else:
                st.write(f"{file_name} -- ⛔ Ce fichier n'est pas un pdf")
    
    # Extraire les informations des PDFs
    if pdf_files:
        with st.spinner("Extraction des informations..."):
            donnees = []
            with zipfile.ZipFile(BytesIO(uploaded_file.getvalue())) as zip_file:
                for pdf_file in pdf_files:
                    pdf_bytes = zip_file.read(pdf_file)
                    lignes = extraire_info_pdf(pdf_bytes)
                    if lignes:
                        for ligne in lignes:
                            ligne['Nom du fichier'] = pdf_file
                            donnees.append(ligne)
            
            time.sleep(1)
        
        st.success("Extraction terminée !")
        
        # Afficher le tableau avec les données
        if donnees:
            st.write("**Informations extraites des bordereaux :**")
            df = pd.DataFrame(donnees)
            # Réorganiser les colonnes
            colonnes = ['Nom du fichier', 'Type', 'Date du paiement', 'Matricule', 'Nom du bénéficiaire', 
                       'Date du', 'Date de fin', 'Nature de la prestation', 'Quantité', 
                       'Montant brut', 'Montant net']
            df = df[colonnes]
            
            # Convertir toutes les colonnes qui doivent être numériques
            df['Quantité'] = df['Quantité'].astype(str).astype(int)
            df['Montant brut'] = df['Montant brut'].astype(str)
            df['Montant net'] = df['Montant net'].astype(str)
            
            # Appliquer un style : fond bleu clair pour les lignes "Montant net"
            def colorer_lignes(row):
                if row['Type'] == 'Montant net':
                    return ['background-color: lightblue'] * len(row)
                return [''] * len(row)
            
            df_styled = df.style.apply(colorer_lignes, axis=1)
            st.dataframe(df_styled, width='stretch', height=400)
            
            # Bouton de téléchargement Excel
            # Créer le fichier Excel en mémoire avec formatage
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Bordereaux')
                
                # Accéder à la feuille de calcul pour ajouter le formatage
                workbook = writer.book
                worksheet = writer.sheets['Bordereaux']
                
                # Importer les styles openpyxl
                from openpyxl.styles import PatternFill
                from datetime import datetime
                
                # Activer les filtres sur la première ligne (entêtes)
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Définir le fond bleu clair
                blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                
                # Identifier les colonnes de dates et de montants
                date_columns = ['Date du paiement', 'Date du', 'Date de fin']
                date_col_indices = [df.columns.get_loc(col) + 1 for col in date_columns if col in df.columns]
                
                montant_columns = ['Montant brut', 'Montant net']
                montant_col_indices = [df.columns.get_loc(col) + 1 for col in montant_columns if col in df.columns]
                
                # Parcourir les lignes du DataFrame et appliquer le style
                for idx, row in df.iterrows():
                    excel_row = idx + 2  # +1 pour l'en-tête, +1 pour l'index Excel qui commence à 1
                    
                    # Appliquer le fond bleu pour les lignes "Montant net"
                    if row['Type'] == 'Montant net':
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=excel_row, column=col).fill = blue_fill
                    
                    # Formater les cellules de dates
                    for col_idx in date_col_indices:
                        cell = worksheet.cell(row=excel_row, column=col_idx)
                        cell_value = cell.value
                        
                        # Convertir la chaîne de date en objet datetime si ce n'est pas vide
                        if cell_value and isinstance(cell_value, str) and cell_value.strip():
                            try:
                                # Format attendu: jj/mm/aaaa
                                date_obj = datetime.strptime(cell_value, '%d/%m/%Y')
                                cell.value = date_obj
                                cell.number_format = 'DD/MM/YYYY'
                            except:
                                pass  # Garder la valeur originale si la conversion échoue
                    
                    # Formater les cellules de montants en numérique avec 2 décimales
                    for col_idx in montant_col_indices:
                        cell = worksheet.cell(row=excel_row, column=col_idx)
                        cell_value = cell.value
                        
                        # Convertir la chaîne de montant en nombre si ce n'est pas vide ou 0
                        if cell_value and isinstance(cell_value, str) and cell_value.strip() and cell_value != '0':
                            try:
                                # Remplacer la virgule par un point pour la conversion
                                montant_float = float(cell_value.replace(',', '.'))
                                cell.value = montant_float
                                cell.number_format = '#,##0.00'
                            except:
                                pass  # Garder la valeur originale si la conversion échoue
                        elif cell_value == '0' or cell_value == 0:
                            cell.value = 0
                            cell.number_format = '#,##0.00'
            
            output.seek(0)
            
            # Bouton de téléchargement
            st.download_button(
                label="Télécharger le fichier Excel",
                data=output,
                file_name="bordereaux_arrêts_maladies.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("Aucune donnée n'a pu être extraite des PDF.")
    else:
        st.warning("Aucun fichier PDF trouvé dans le dossier.")